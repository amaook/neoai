import csv
import datetime as _dt
import json
import os
import re
import sys
from decimal import Decimal, InvalidOperation


def is_empty(value):
    return value is None or (isinstance(value, str) and value.strip() == "")


def normalize_cell(value):
    if isinstance(value, str):
        return value.strip()
    return value


def detect_encoding(path):
    try:
        import charset_normalizer

        with open(path, "rb") as fh:
            raw = fh.read(65536)
        best = charset_normalizer.from_bytes(raw).best()
        if best and best.encoding:
            return best.encoding
    except Exception:
        pass
    for encoding in ("utf-8-sig", "utf-8", "gb18030", "gbk", "big5", "latin1"):
        try:
            with open(path, "r", encoding=encoding) as fh:
                fh.read(4096)
            return encoding
        except Exception:
            continue
    return "utf-8-sig"


def read_csv_rows(path):
    ext = os.path.splitext(path)[1].lower()
    delimiter = "\t" if ext == ".tsv" else ","
    encoding = detect_encoding(path)
    with open(path, "r", encoding=encoding, newline="") as fh:
        return [row for row in csv.reader(fh, delimiter=delimiter)], encoding


def read_excel_rows(path, sheet_name):
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=False)
    names = [sheet_name] if sheet_name else workbook.sheetnames
    sheets = []
    for name in names:
        if name not in workbook.sheetnames:
            raise ValueError("找不到工作表：" + name)
        ws = workbook[name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
        sheets.append({"name": name, "rows": rows})
    return sheets


def write_xlsx(path, sheets):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    workbook.remove(workbook.active)
    for index, sheet in enumerate(sheets):
        title = re.sub(r"[\[\]\*:/\\?]", "_", str(sheet["name"] or f"Sheet{index + 1}"))[:31] or f"Sheet{index + 1}"
        if title in workbook.sheetnames:
            title = f"{title[:26]}_{index + 1}"
        ws = workbook.create_sheet(title)
        rows = sheet.get("rows") or []
        for row in rows:
            ws.append(row)
        if rows:
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="F5F4F0")
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            column_count = max(len(row) for row in rows)
            for col_index in range(1, column_count + 1):
                max_len = 10
                for row in rows[:300]:
                    if col_index <= len(row):
                        max_len = max(max_len, min(42, len(str(row[col_index - 1] or "")) + 2))
                ws.column_dimensions[get_column_letter(col_index)].width = max_len
    workbook.save(path)


def write_delimited(path, rows):
    ext = os.path.splitext(path)[1].lower()
    delimiter = "\t" if ext == ".tsv" else ","
    with open(path, "w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.writer(fh, delimiter=delimiter)
        writer.writerows(rows)


def pad_rows(rows):
    width = max([len(row) for row in rows] + [0])
    return [list(row) + [None] * (width - len(row)) for row in rows]


def header_index(header):
    return {str(name).strip(): idx for idx, name in enumerate(header) if not is_empty(name)}


def selected_indexes(header, columns):
    index = header_index(header)
    return [index[col] for col in columns if col in index]


def auto_indexes(header, kind):
    if kind == "amount":
        tokens = ("金额", "收入", "支出", "费用", "成本", "利润", "价格", "单价", "总价", "余额", "amount", "price", "revenue", "cost", "profit", "fee", "money")
    elif kind == "date":
        tokens = ("日期", "时间", "年月", "月份", "date", "time", "day", "month")
    else:
        tokens = ()
    result = []
    for idx, name in enumerate(header):
        text = str(name or "").strip().lower()
        if any(token.lower() in text for token in tokens):
            result.append(idx)
    return result


def coerce_number(value):
    if is_empty(value):
        return None
    if isinstance(value, (int, float, Decimal)):
        return value
    text = str(value).strip()
    text = re.sub(r"[,\s￥¥$]", "", text)
    text = text.replace("(", "-").replace(")", "")
    try:
        return float(Decimal(text))
    except (InvalidOperation, ValueError):
        return value


def coerce_date(value):
    if is_empty(value):
        return value
    if isinstance(value, (_dt.date, _dt.datetime)):
        return value
    text = str(value).strip().replace(".", "-").replace("/", "-")
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%Y%m%d", "%m-%d-%Y"):
        try:
            return _dt.datetime.strptime(text, fmt)
        except ValueError:
            pass
    return value


def compare_value(left, op, right):
    if op == "is_empty":
        return is_empty(left)
    if op == "not_empty":
        return not is_empty(left)
    if op in ("contains", "not_contains"):
        result = str(right) in str(left or "")
        return result if op == "contains" else not result
    if op in ("eq", "ne"):
        result = str(left) == str(right)
        return result if op == "eq" else not result
    left_num = coerce_number(left)
    right_num = coerce_number(right)
    if isinstance(left_num, (int, float)) and isinstance(right_num, (int, float)):
        values = (left_num, right_num)
    else:
        values = (str(left or ""), str(right or ""))
    if op == "gt":
        return values[0] > values[1]
    if op == "gte":
        return values[0] >= values[1]
    if op == "lt":
        return values[0] < values[1]
    if op == "lte":
        return values[0] <= values[1]
    return True


def apply_operation(rows, operation):
    op_type = operation.get("type")
    rows = pad_rows(rows)
    if not rows:
        return rows
    header = rows[0]
    body = rows[1:]

    if op_type == "drop_empty_rows":
        return [row for row in rows if any(not is_empty(cell) for cell in row)]

    if op_type == "drop_empty_columns":
        keep = []
        width = max([len(row) for row in rows] + [0])
        for idx in range(width):
            if any(idx < len(row) and not is_empty(row[idx]) for row in rows):
                keep.append(idx)
        return [[row[idx] if idx < len(row) else None for idx in keep] for row in rows]

    if op_type == "trim_text":
        cols = operation.get("columns") or []
        target = selected_indexes(header, cols) if cols else list(range(len(header)))
        for row in rows:
            for idx in target:
                if idx < len(row) and isinstance(row[idx], str):
                    row[idx] = row[idx].strip()
        return rows

    if op_type == "drop_duplicates":
        cols = operation.get("columns") or []
        target = selected_indexes(header, cols) if cols else list(range(len(header)))
        seen = set()
        cleaned = [header]
        for row in body:
            key = tuple(row[idx] if idx < len(row) else None for idx in target)
            if key in seen:
                continue
            seen.add(key)
            cleaned.append(row)
        return cleaned

    if op_type == "fill_empty":
        cols = operation.get("columns") or []
        value = operation.get("value", "")
        target = selected_indexes(header, cols) if cols else list(range(len(header)))
        for row in body:
            for idx in target:
                if idx < len(row) and is_empty(row[idx]):
                    row[idx] = value
        return [header] + body

    if op_type == "rename_columns":
        mapping = operation.get("mapping") or {}
        return [[mapping.get(str(cell), cell) for cell in header]] + body

    if op_type == "select_columns":
        cols = operation.get("columns") or []
        target = selected_indexes(header, cols)
        return [[row[idx] if idx < len(row) else None for idx in target] for row in rows]

    if op_type == "remove_columns":
        cols = set(operation.get("columns") or [])
        target = [idx for idx, name in enumerate(header) if str(name).strip() not in cols]
        return [[row[idx] if idx < len(row) else None for idx in target] for row in rows]

    if op_type == "filter_rows":
        column = operation.get("column")
        idx = header_index(header).get(column)
        if idx is None:
            return rows
        op = operation.get("op") or "eq"
        value = operation.get("value")
        return [header] + [row for row in body if compare_value(row[idx] if idx < len(row) else None, op, value)]

    if op_type == "normalize_amount":
        cols = operation.get("columns") or []
        target = selected_indexes(header, cols) if cols else auto_indexes(header, "amount")
        for row in body:
            for idx in target:
                row[idx] = coerce_number(row[idx] if idx < len(row) else None)
        return [header] + body

    if op_type == "normalize_date":
        cols = operation.get("columns") or []
        target = selected_indexes(header, cols) if cols else auto_indexes(header, "date")
        for row in body:
            for idx in target:
                row[idx] = coerce_date(row[idx] if idx < len(row) else None)
        return [header] + body

    if op_type == "sort":
        cols = operation.get("by") or operation.get("columns") or []
        target = selected_indexes(header, cols)
        reverse = operation.get("ascending") is False
        if target:
            body = sorted(body, key=lambda row: tuple(row[idx] if idx < len(row) else None for idx in target), reverse=reverse)
        return [header] + body

    return rows


def default_operations(options):
    ops = []
    if options.get("drop_empty_rows", True):
        ops.append({"type": "drop_empty_rows"})
    if options.get("drop_empty_columns", True):
        ops.append({"type": "drop_empty_columns"})
    if options.get("trim_text", True):
        ops.append({"type": "trim_text"})
    if options.get("drop_duplicates", True):
        ops.append({"type": "drop_duplicates", "columns": options.get("duplicate_columns") or []})
    if options.get("normalize_amount", True):
        ops.append({"type": "normalize_amount", "columns": options.get("amount_columns") or []})
    if options.get("normalize_date", True):
        ops.append({"type": "normalize_date", "columns": options.get("date_columns") or []})
    if "fill_empty" in options:
        ops.append({"type": "fill_empty", "value": options.get("fill_empty")})
    return ops


def clean_rows(rows, operations, options):
    cleaned = pad_rows(rows)
    ops = operations or default_operations(options or {})
    before_rows = len(cleaned)
    before_cols = max([len(row) for row in cleaned] + [0])
    for operation in ops:
        cleaned = apply_operation(cleaned, operation or {})
    after_rows = len(cleaned)
    after_cols = max([len(row) for row in cleaned] + [0])
    return cleaned, {
        "beforeRows": before_rows,
        "beforeColumns": before_cols,
        "afterRows": after_rows,
        "afterColumns": after_cols,
        "operations": [op.get("type") for op in ops if op.get("type")],
    }


def main():
    payload = json.loads(sys.argv[1])
    input_path = payload["input_path"]
    output_path = payload["output_path"]
    sheet_name = payload.get("sheet") or ""
    operations = payload.get("operations") or []
    options = payload.get("options") or {}
    ext = os.path.splitext(input_path)[1].lower()

    if ext in (".csv", ".tsv"):
        rows, encoding = read_csv_rows(input_path)
        cleaned, summary = clean_rows(rows, operations, options)
        sheets = [{"name": "Cleaned", "rows": cleaned}]
        summary["encoding"] = encoding
    elif ext in (".xlsx", ".xlsm"):
        sheets = []
        summary = {"sheets": []}
        for sheet in read_excel_rows(input_path, sheet_name):
            cleaned, sheet_summary = clean_rows(sheet["rows"], operations, options)
            sheets.append({"name": sheet["name"], "rows": cleaned})
            summary["sheets"].append({"name": sheet["name"], **sheet_summary})
    else:
        raise ValueError("不支持的表格格式：" + ext)

    out_ext = os.path.splitext(output_path)[1].lower()
    if out_ext in (".csv", ".tsv"):
        write_delimited(output_path, sheets[0]["rows"] if sheets else [])
    else:
        write_xlsx(output_path, sheets)

    print(json.dumps({
        "ok": True,
        "summary": summary,
        "outputFormat": out_ext.lstrip(".") or "xlsx"
    }, ensure_ascii=False))


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(json.dumps({"ok": False, "error": str(exc)}, ensure_ascii=False))
        sys.exit(1)
